import os
import time
import pandas as pd
import openpyxl
import shutil
import secrets
import threading
from flask import Flask, request, render_template, Response, send_from_directory,jsonify,make_response,session
from flask_session import Session
import langsmith.client
from langgraph.graph import END, StateGraph, START
from langgraph.utils.runnable import Runnable
from langsmith import trace
from typing import TypedDict, Sequence
from langchain_core.tracers.context import tracing_v2_enabled

# Directly configure LangSmith API
LANGSMITH_TRACING = True
LANGSMITH_ENDPOINT = "https://api.smith.langchain.com"
LANGSMITH_API_KEY = "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"

# Import functions from other scripts
from Contract_Rules_Extraction_new import extract_contract_rules
from new_data_extraction import process_pdfs_in_directory_with_azure
from new_data_validation import validate_all_data

app = Flask(__name__, template_folder="templates")
app.secret_key = secrets.token_hex(16)
# Configure server-side session storage
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"  # Store sessions in files
Session(app)


# Configure upload folders
INVOICE_UPLOAD_FOLDER = "/Users/venkatasaimaruthi.n/Downloads/makeathon_code/inputs"
GROUNDING_DATASET_UPLOAD_FOLDER="/Users/venkatasaimaruthi.n/Downloads/makeathon_code/grounding-datasets/preprocessed-grounding-dataset"
SERVICE_AGREEMENT_UPLOAD_FOLDER = "/Users/venkatasaimaruthi.n/Downloads/makeathon_code/rules"
OUTPUT_DIRECTORY = "/Users/venkatasaimaruthi.n/Downloads/makeathon_code/output/validation_reports"
EXTRACTED_RULES_FILE = "/Users/venkatasaimaruthi.n/Downloads/makeathon_code/rules/extracted_rules.txt"

os.makedirs(INVOICE_UPLOAD_FOLDER, exist_ok=True)
os.makedirs(GROUNDING_DATASET_UPLOAD_FOLDER, exist_ok=True)
os.makedirs(SERVICE_AGREEMENT_UPLOAD_FOLDER, exist_ok=True)

app.config["INVOICE_UPLOAD_FOLDER"] = INVOICE_UPLOAD_FOLDER
app.config["SERVICE_AGREEMENT_UPLOAD_FOLDER"] = SERVICE_AGREEMENT_UPLOAD_FOLDER
app.config["GROUNDING_DATASET_UPLOAD_FOLDER"] = GROUNDING_DATASET_UPLOAD_FOLDER

status_updates = []
status_lock = threading.Lock()
ready_for_download = False

# Define the state globally
global_state = {
    "user_decision": None
}

def update_status(message):
    global ready_for_download
    with status_lock:
        status_updates.append(message)
        print(f"Status update: {message}")
        if message == "Reports are ready for download.":
            ready_for_download = True

@app.route("/status")
def status():
    def stream():
        last_update = 0
        while True:
            with status_lock:
                if last_update < len(status_updates):
                    yield f"data: {status_updates[last_update]}\n\n"
                    last_update += 1
    return Response(stream(), content_type="text/event-stream")

@app.route("/is_ready", methods=["GET"])
def is_ready():
    global ready_for_download
    return jsonify({"ready": ready_for_download})

@app.route("/get_agreement")
def get_agreement():
    name = request.args.get("name")
    file_path = os.path.join(SERVICE_AGREEMENT_UPLOAD_FOLDER, name)

    if os.path.exists(file_path):
        with open(file_path, "r", encoding="utf-8") as file:
            return file.read()
    return "Agreement not found", 404

@app.route("/save_agreement", methods=["POST"])
def save_agreement():
    data = request.json
    name = data.get("name")
    content = data.get("content")

    if not name or not content:
        return jsonify({"message": "Invalid request"}), 400

    file_path = os.path.join(SERVICE_AGREEMENT_UPLOAD_FOLDER, name)
    with open(file_path, "w", encoding="utf-8") as file:
        file.write(content)

    # Update the global state with user decision and service agreement details
    global global_state
    global_state["user_decision"] = "yes"
    global_state["service_agreement"] = file_path

    # Include invoices and grounding datasets in the state
    state = {
        "invoices": global_state.get("invoices", []),
        "grounding_datasets": global_state.get("grounding_datasets", []),
        "service_agreement": global_state.get("service_agreement", "")
    }

    # Trigger the continuation of the workflow
    run_partial_analysis(state)

    return jsonify({"message": "Agreement saved and workflow resumed."})

@app.route("/save_existing_agreement", methods=["POST"])
def save_existing_agreement():
    data = request.json
    name = data.get("name")
    content = data.get("content")

    if not name or not content:
        return jsonify({"success": False, "message": "Invalid request"}), 400

    # Save the agreement content
    file_path = os.path.join(SERVICE_AGREEMENT_UPLOAD_FOLDER, name)
    try:
        with open(file_path, "w", encoding="utf-8") as file:
            file.write(content)

        # Retrieve the global state from the session
        global_state = session.get('global_state', {})

        # Log session state to verify its contents
        print("Session State:", global_state)

        global_state["service_agreement"] = file_path

        
        # Prepare the state for analysis
        state = {
            "invoices": global_state.get("invoices", []),
            "grounding_datasets": global_state.get("grounding_datasets", []),
            "service_agreement": global_state.get("service_agreement", "")
        }

        # Log global_state to verify its contents
        print("Global State before running partial analysis:", state)

        # Trigger partial analysis
        run_partial_analysis(state)

        return jsonify({"success": True, "message": "Agreement saved and analysis triggered."})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/edit_rules", methods=["GET"])
def edit_rules():
    name = "extracted_rules.txt"
    file_path = os.path.join(SERVICE_AGREEMENT_UPLOAD_FOLDER, name)

    if os.path.exists(file_path):
        with open(file_path, "r", encoding="utf-8") as file:
            return file.read()
    return "Rules file not found.", 404 

@app.route("/user_decision", methods=["POST"])
def user_decision():
    data = request.json
    decision = data.get("decision")
    if decision in ["yes", "no"]:
        global global_state
        global_state["user_decision"] = decision
        return jsonify({"message": "User decision recorded."})
    return jsonify({"message": "Invalid decision."}), 400


@app.route("/", methods=["GET", "POST"])
def upload_files():
    # List files in the service agreement folder (for "Existing" file selection)
    service_agreement_folder = SERVICE_AGREEMENT_UPLOAD_FOLDER
    existing_rules = []
    
    # Only list .txt files from the folder
    if os.path.exists(service_agreement_folder):
        existing_rules = [f for f in os.listdir(service_agreement_folder) if f.endswith('.txt')]

    file_exists = os.path.exists(EXTRACTED_RULES_FILE)
    
    if request.method == "POST":
        update_status("File upload initiated.")
        metadata = {}

        # Handle multiple invoices
        invoices = request.files.getlist("invoice")
        invoice_paths = []
        for invoice in invoices:
            if invoice.filename:
                invoice_path = os.path.join(INVOICE_UPLOAD_FOLDER, invoice.filename)
                invoice.save(invoice_path)
                invoice_paths.append(invoice_path)
                update_status(f"Invoice '{invoice.filename}' uploaded.")
        metadata["invoices"] = invoice_paths

        # Handle multiple grounding-datasets
        grounding_datasets = request.files.getlist("grounding_dataset")
        grounding_dataset_paths = []
        for grounding_dataset in grounding_datasets:
            if grounding_dataset.filename:
                grounding_dataset_path = os.path.join(GROUNDING_DATASET_UPLOAD_FOLDER, grounding_dataset.filename)
                
                try:
                    grounding_dataset.save(grounding_dataset_path)
                    grounding_dataset_paths.append(grounding_dataset_path)
                    print(f"File saved successfully: {grounding_dataset_path}")
                except Exception as e:
                    print(f"Error saving {grounding_dataset.filename}: {e}")
        
                update_status(f"Grounding dataset '{grounding_dataset.filename}' uploaded.")

        metadata["grounding_datasets"] = grounding_dataset_paths

        service_agreement_option = request.form.get("service_agreement_option")
        metadata["service_agreement_option"] = service_agreement_option

        if service_agreement_option == "new":
            service_agreement = request.files["service_agreement"]
            if service_agreement.filename:
                agreement_path = os.path.join(SERVICE_AGREEMENT_UPLOAD_FOLDER, service_agreement.filename)
                service_agreement.save(agreement_path)
                metadata["service_agreement"] = agreement_path
                update_status("New Service Agreement uploaded.")
            else:
                metadata["service_agreement"] = None
        else:  # Existing selected
            selected_rule = request.form.get("existing_rule")
            metadata["service_agreement"] = selected_rule
            update_status(f"Using existing service agreement rules: {selected_rule}")

        update_status("Starting analysis workflow...")
        time.sleep(2)

        # Update global state with uploaded files
        global global_state
        global_state["invoices"] = metadata.get("invoices", [])
        global_state["grounding_datasets"] = metadata.get("grounding_datasets", [])
        global_state["service_agreement"] = metadata["service_agreement"]

        # Store the state in a session variable
        session['global_state'] = global_state
        
        # Log global_state to verify its contents
        print("Global State after file upload:", global_state)

        run_analysis(metadata)

    return render_template("upload_1.html", file_exists=file_exists, existing_rules=existing_rules)

@app.route("/download_report")
def download_report():
    zip_file_path = "/Users/venkatasaimaruthi.n/Downloads/makeathon_code/output/validation_reports.zip"
    
    if os.path.exists(OUTPUT_DIRECTORY):
        # Create the ZIP archive if it doesn't exist
        shutil.make_archive(zip_file_path.replace(".zip", ""), "zip", OUTPUT_DIRECTORY)
        
        # Use send_from_directory to serve the file
        response = make_response(
            send_from_directory(
                directory=os.path.dirname(zip_file_path),
                path=os.path.basename(zip_file_path),
                as_attachment=True
            )
        )
        
        # Add Cache-Control header to prevent caching
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"  # For older HTTP/1.0 clients
        response.headers["Expires"] = "0"       # Forces the file to expire immediately
        return response
    else:
        update_status("Validation report not found.")
        return "Validation report not found.", 404

def run_analysis(state):
    update_status("Initializing analysis workflow...")
    time.sleep(2)

    class State(TypedDict):
        messages: str
        invoices: list
        grounding_datasets: list
        service_agreement_option: str

    builder = StateGraph(State)
    
    # Node to check service agreement type
    def check_service_agreement(state):
        update_status("Checking service agreement type...")
        time.sleep(1)
        return {"service_agreement_option": state["service_agreement_option"]}
        

    # Node for extracting contract rules
    def extract_contract_rules_with_status(state):
        update_status("Extracting contract rules from new Service Agreement...")
        result = extract_contract_rules(state)
        time.sleep(2)
        update_status("Contract rules extracted successfully.")
        return result
    
    def await_user_decision(state):
        update_status("Do you want to view/edit the generated rules?")
        global global_state
        
        # Pause execution and wait for user input via frontend
        while not global_state.get("user_decision"):
            time.sleep(1)
        
        decision = global_state["user_decision"]
        if decision == "yes":
            update_status("Loading the generated rules to editable area.")
            extracted_rules_path = os.path.join(SERVICE_AGREEMENT_UPLOAD_FOLDER, "extracted_rules_new.txt")
            if os.path.exists(extracted_rules_path):
                with open(extracted_rules_path, "r", encoding="utf-8") as file:
                    state["extracted_rules"] = file.read()
            return state
        elif decision == "no":
            update_status("User decided not to view/edit the rules. Proceeding with data extraction.")
            return state
     
    builder.add_node("check_service_agreement", check_service_agreement)
    builder.add_node("await_user_decision", await_user_decision)
    builder.add_node("rules_extractor", extract_contract_rules_with_status)
    builder.add_node("data_extraction", data_extraction_task)
    builder.add_node("pre_processing", pre_processing_task)
    builder.add_node("file_comparison", file_comparison_task)

    def validate_all_data_with_status(state):
        update_status("Starting data validation against contract rules...")
        time.sleep(60)
        result = validate_all_data(state)
        update_status("Data validation completed.")
        time.sleep(2)
        update_status("Reports are ready for download.")
        return result

    builder.add_node("validation", validate_all_data_with_status)

    # Define the routing logic
    def route_service_agreement(state: State) -> Sequence[str]:
        if state["service_agreement_option"] == "new":
            return ["rules_extractor"]
        return ["data_extraction"]
    
    def route_user_decision(state: State) -> Sequence[str]:
        decision = global_state.get("user_decision", "")
        if decision == "yes":
            extracted_rules_path = os.path.join(SERVICE_AGREEMENT_UPLOAD_FOLDER, "extracted_rules_new.txt")
            if os.path.exists(extracted_rules_path):
                with open(extracted_rules_path, "r", encoding="utf-8") as file:
                    state["extracted_rules"] = file.read()
            return [END]
        elif decision == "no":
            return ["data_extraction"]
        return ["data_extraction"]
    
    builder.add_edge(START, "check_service_agreement")
    
    # Add conditional edges
    builder.add_conditional_edges(
        "check_service_agreement",
        route_service_agreement,
        ["rules_extractor", "data_extraction"],
    )

    builder.add_edge("rules_extractor", "await_user_decision")

    builder.add_conditional_edges(
        "await_user_decision",
        route_user_decision,
        ["data_extraction", END]
    )
    
    #builder.add_edge("await_user_decision", "data_extraction")
    builder.add_edge("data_extraction", "pre_processing")
    builder.add_edge("pre_processing", "file_comparison")
    builder.add_edge("file_comparison", "validation")
    builder.add_edge("validation", END)

    graph = builder.compile()
    image=graph.get_graph().draw_mermaid_png()
    with open("graph_image.png","wb") as file:
        file.write(image)


    formatted_state = {
        "messages": state.get("service_agreement", "No service agreement provided"),
        "invoices": state.get("invoices", []),
        "grounding_datasets": state.get("grounding_datasets",[]),
        "service_agreement_option": state.get("service_agreement_option", "new"),
        "user_decision": global_state.get("user_decision", ""),
    }

    langsmith_client = langsmith.Client(api_url="https://api.smith.langchain.com", api_key=LANGSMITH_API_KEY)

    with tracing_v2_enabled(client=langsmith_client):
        graph.invoke(formatted_state)

    #if global_state["user_decision"] == "yes" or global_state["user_decision"] == "no":
        #update_status("Reports are ready for download.")
    #return True


def run_partial_analysis(state):
    update_status("Starting partial analysis workflow from data_extraction...")
    time.sleep(2)

    class State(TypedDict):
        invoices: list
        grounding_datasets: list

    builder = StateGraph(State)

    def validate_all_data_with_status(state):
        update_status("Starting data validation against contract rules...")
        time.sleep(60)
        result = validate_all_data(state)
        update_status("Data validation completed.")
        time.sleep(2)
        update_status("Reports are ready for download.")
        return result

    builder.add_node("data_extraction", data_extraction_task)
    builder.add_node("pre_processing", pre_processing_task)
    builder.add_node("file_comparison", file_comparison_task)
    builder.add_node("validation", validate_all_data_with_status)
    
    builder.add_edge(START, "data_extraction")
    builder.add_edge("data_extraction", "pre_processing")
    builder.add_edge("pre_processing", "file_comparison")
    builder.add_edge("file_comparison", "validation")
    builder.add_edge("validation", END)

    graph = builder.compile()

    formatted_state = {
        "invoices": state.get("invoices", []),
        "grounding_datasets": state.get("grounding_datasets", []),
        "service_agreement": state.get("service_agreement", "")
    }
    
    # Debugging logs to check the formatted_state
    print("Formatted State for Partial Analysis:", formatted_state)

    langsmith_client = langsmith.Client(api_url="https://api.smith.langchain.com", api_key=LANGSMITH_API_KEY)

    with tracing_v2_enabled(client=langsmith_client):
        graph.invoke(formatted_state)


def data_extraction_task(state): 
    update_status("Starting data extraction...")
    invoice_paths = state.get("invoices", [])
    
    if not invoice_paths:
        update_status("No invoices provided. Skipping data extraction.")
        return state

    output_dir_json = "/Users/venkatasaimaruthi.n/Downloads/makeathon_code/output/json"
    output_dir_excels = "/Users/venkatasaimaruthi.n/Downloads/makeathon_code/output/excels"
    azure_endpoint = "https://myformrecog2943.cognitiveservices.azure.com/"
    azure_api_key = "EH8Y3Ao3qJ16yxLH2g74fIO2c42NpwS8UeVXuYyMdTBt2OMr86Z5JQQJ99BAACYeBjFXJ3w3AAALACOG7X6r"

    try:
        process_pdfs_in_directory_with_azure(invoice_paths, output_dir_json,output_dir_excels, azure_endpoint, azure_api_key)
        update_status("Data extraction completed.")
    except Exception as e:
        update_status(f"Error during data extraction: {e}")
    return state

def pre_processing_task(state):
    
    # Retrieve grounding dataset paths from the state
    grounding_dataset_paths = state.get("grounding_datasets", [])
    
    if not grounding_dataset_paths:
        update_status("No grounding datasets uploaded. Skipping preprocessing.")
        return state

    # Define dynamic input directory based on uploaded datasets
    input_directory = os.path.dirname(grounding_dataset_paths[0])  # Use directory of first uploaded file
    output_directory = "/Users/venkatasaimaruthi.n/Downloads/makeathon_code/grounding-datasets/preprocessed-grounding-dataset"

    input_directory_excel = "/Users/venkatasaimaruthi.n/Downloads/makeathon_code/output/excels"
    output_directory_excel = "/Users/venkatasaimaruthi.n/Downloads/makeathon_code/grounding-datasets/preprocessed-excel-output-directory"

    # Create necessary output directories if they don’t exist
    os.makedirs(output_directory, exist_ok=True)
    os.makedirs(output_directory_excel, exist_ok=True)

    update_status(f"Preprocessing data....")

    # Function to format columns with "date" in their name
    def format_date_columns(df):
        for column in df.columns:
            if "date" in column.lower():
                try:
                    df[column] = pd.to_datetime(df[column], errors='coerce')
                    df[column] = df[column].dt.strftime('%m/%d/%Y')
                except Exception:
                    continue
        return df

    # Function to clean data
    def clean_data(df):
        df.columns = df.columns.str.lower().str.replace(r'[\n\r\s]+', '', regex=True)
        df = df.applymap(lambda x: str(x).replace('$', '').strip() if isinstance(x, str) else x)
        if 'contractorname' in df.columns:
            df['contractorname'] = df['contractorname'].fillna(method='ffill')
        if 'role' in df.columns:
            df['role'] = df['role'].fillna(method='ffill')
        return df

    # Function to remove rows containing "total"
    def remove_total_rows(df):
        return df[~df.apply(lambda row: row.astype(str).str.contains('total', case=False).any(), axis=1)]

    # Process each Excel file in the input directory
    for file_name in os.listdir(input_directory_excel):
        if file_name.endswith((".xlsx", ".xls")):
            input_file_path = os.path.join(input_directory_excel, file_name)
            output_file_path = os.path.join(output_directory_excel, file_name)

            try:
                excel_file = pd.ExcelFile(input_file_path)
                visible_sheets = [sheet for sheet in excel_file.sheet_names if not excel_file.parse(sheet).empty]

                if len(visible_sheets) >= 3:
                    third_sheet_name = visible_sheets[2]
                    third_sheet_df = pd.read_excel(input_file_path, sheet_name=third_sheet_name)

                    third_sheet_df = remove_total_rows(third_sheet_df)
                    cleaned_df = clean_data(third_sheet_df)

                    output_csv_file_path = os.path.join(output_directory_excel, f"{os.path.splitext(file_name)[0]}.csv")
                    cleaned_df.to_csv(output_csv_file_path, index=False)
                else:
                    print(f"Skipped: {input_file_path} (less than 3 visible sheets)")

            except Exception as e:
                print(f"Error processing {input_file_path}: {e}")
    
    update_status(f"Pre-processing completed")
    return state

def file_comparison_task(state):
    import os
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    # Directories
    new_csv_dir_1 = state.get("new_csv_dir_1", "/Users/venkatasaimaruthi.n/Downloads/makeathon_code/grounding-datasets/preprocessed-grounding-dataset")
    new_csv_dir_2 = state.get("new_csv_dir_2", "/Users/venkatasaimaruthi.n/Downloads/makeathon_code/grounding-datasets/preprocessed-excel-output-directory")
    output_comparison_dir = "/Users/venkatasaimaruthi.n/Downloads/makeathon_code/output/validation_reports"
    update_status("Started comparing Extracted Data vs Grounded dataset")

    # Highlighting function for changes
    def highlight_changes(ws, row, col):
        fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        ws.cell(row=row + 1, column=col + 1).fill = fill  # Adjust for 1-based indexing

    # Highlighting function for matches
    def highlight_matches(ws, row, col):
        fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        ws.cell(row=row + 1, column=col + 1).fill = fill  # Adjust for 1-based indexing

    # Function to standardize and clean data
    def clean_data(value, column_name):
        if isinstance(value, str):
            if column_name == 'description':
                value = value.replace("\n", " ").replace("\r", " ")  # Remove line breaks but keep hyphens
            value = value.strip().replace(" ", "")
            return value
        else:
            return str(value).strip().replace(" ", "")

    # New logic to highlight missing columns in yellow
    def highlight_missing_columns(ws, missing_columns, max_row):
        fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red fill
        col_idx = ws.max_column + 1  # Start after the existing columns

        for col_name in missing_columns:
            ws.cell(row=1, column=col_idx, value=col_name)  # Add column name to header
            for row_idx in range(2, max_row + 1):  # Highlight all rows in the column
                ws.cell(row=row_idx, column=col_idx).fill = fill
            col_idx += 1

    # Compare CSV files from two directories
    for file_name in os.listdir(new_csv_dir_1):
        if file_name.endswith(".csv"):
            csv_file_path_1 = os.path.join(new_csv_dir_1, file_name)
            csv_file_path_2 = os.path.join(new_csv_dir_2, file_name)
            output_file_path = os.path.join(output_comparison_dir, file_name.replace(".csv", "_comparison.xlsx"))

            if os.path.exists(csv_file_path_2):
                try:
                    # Load CSV files
                    csv_df_1 = pd.read_csv(csv_file_path_1, header=0)
                    csv_df_2 = pd.read_csv(csv_file_path_2, header=0)

                    if csv_df_1.empty or csv_df_2.empty:
                        print(f"Warning: One of the files ({file_name}) is empty.")
                        continue

                    # Standardize column names
                    csv_df_1.columns = csv_df_1.columns.str.strip().str.replace(" ", "").str.lower()
                    csv_df_2.columns = csv_df_2.columns.str.strip().str.replace(" ", "").str.lower()

                    # Identify missing columns
                    missing_in_1 = set(csv_df_2.columns) - set(csv_df_1.columns)
                    missing_in_2 = set(csv_df_1.columns) - set(csv_df_2.columns)

                    if missing_in_1:
                        print(f"Columns missing in {file_name} (Directory 1): {missing_in_1}.")
                    if missing_in_2:
                        print(f"Columns missing in {file_name} (Directory 2): {missing_in_2}.")

                    # Apply cleaning to both DataFrames
                    for col in csv_df_1.columns:
                        csv_df_1[col] = csv_df_1[col].apply(lambda x: clean_data(x, col))
                    for col in csv_df_2.columns:
                        csv_df_2[col] = csv_df_2[col].apply(lambda x: clean_data(x, col))

                    # Convert specific columns to integers (rate, amount, hours)
                    columns_to_convert = ['rate', 'amount', 'hours']
                    for col in columns_to_convert:
                        if col in csv_df_1.columns:
                            csv_df_1[col] = pd.to_numeric(csv_df_1[col], errors='coerce').fillna(0).astype(int)
                        if col in csv_df_2.columns:
                            csv_df_2[col] = pd.to_numeric(csv_df_2[col], errors='coerce').fillna(0).astype(int)

                    # Standardize date columns to ignore time and timezones
                    date_columns = ['date']
                    for col in date_columns:
                        if col in csv_df_1.columns:
                            csv_df_1[col] = pd.to_datetime(csv_df_1[col], errors='coerce').dt.date
                        if col in csv_df_2.columns:
                            csv_df_2[col] = pd.to_datetime(csv_df_2[col], errors='coerce').dt.date

                    # Identify matching columns by name
                    common_columns = list(set(csv_df_1.columns) & set(csv_df_2.columns))

                    if not common_columns:
                        print(f"No common columns found in {file_name}. Skipping comparison.")
                        continue

                    csv_df_1 = csv_df_1[common_columns]
                    csv_df_2 = csv_df_2[common_columns]

                    # Initialize list to track changes
                    differences = []
                    matches = []
                    min_rows = min(len(csv_df_1), len(csv_df_2))

                    # Compare row by row, cell by cell
                    for row_idx in range(min_rows):
                        for col_idx, col in enumerate(common_columns):
                            val_1 = csv_df_1.iloc[row_idx, col_idx]
                            val_2 = csv_df_2.iloc[row_idx, col_idx]
                            if val_1 != val_2:
                                differences.append((row_idx, col_idx))
                            else:
                                matches.append((row_idx, col_idx))

                    # Save differences as a highlighted Excel file
                    with pd.ExcelWriter(output_file_path, engine="openpyxl") as writer:
                        csv_df_1.to_excel(writer, sheet_name="grounding_dataset_data", index=False)
                        csv_df_2.to_excel(writer, sheet_name="output_data", index=False)
                        wb = writer.book
                        ws1 = wb["grounding_dataset_data"]
                        ws2 = wb["output_data"]

                        # Highlight changes in cells
                        for row_idx, col_idx in differences:
                            highlight_changes(ws1, row_idx, col_idx)
                            highlight_changes(ws2, row_idx, col_idx)

                        # Highlight matches in cells
                        for row_idx, col_idx in matches:
                            highlight_matches(ws1, row_idx, col_idx)
                            highlight_matches(ws2, row_idx, col_idx)

                        # Highlight missing columns in Directory 1
                        if missing_in_1:
                            highlight_missing_columns(ws1, missing_in_1, ws1.max_row)

                        # Highlight missing columns in Directory 2
                        if missing_in_2:
                            highlight_missing_columns(ws2, missing_in_2, ws2.max_row)

                    print(f"Differences highlighted in: {output_file_path}")

                except Exception as e:
                    print(f"Error processing {file_name}: {e}")
    update_status("Data comparison completed.")
    


if __name__ == "__main__":
    app.run(debug=True)
